"""
AI-Powered Data Normalizer
Uses Gemini to parse and normalize uploaded financial data.
All AI decisions are logged in the audit trail for regulatory compliance.
"""
import pandas as pd
import openpyxl
import io
import json
from typing import Optional
from datetime import datetime
from loguru import logger

from core.gemini_client import GeminiClient
from core.audit_trail import AuditRecord
from core.schemas import (
    GeneralLedger, JournalEntry, ChartOfAccounts, Account, 
    TrialBalance, TrialBalanceRow
)

# ---------------------------------------------------------------------------
# Constants for chunked / large-file processing
# ---------------------------------------------------------------------------
CHUNK_SIZE = 50_000       # rows per processing chunk for large files
MAX_ROWS_TOTAL = 1_000_000  # hard cap - reject files larger than this
SAMPLE_ROWS = 15          # rows to sample per sheet for AI detection


class DataNormalizer:
    """
    AI-powered normalizer for uploaded financial data.
    Uses Gemini to intelligently parse and normalize data from any format.
    All parsing decisions are logged for audit trail.
    """
    
    def __init__(self):
        self.gemini = GeminiClient()
    
    async def parse_file(
        self,
        content: bytes,
        filename: str,
        file_type: str,
        audit_record: Optional[AuditRecord] = None
    ):
        """
        Parse uploaded file using AI-powered normalization.
        
        Args:
            content: File content as bytes
            filename: Original filename
            file_type: Type of file (general_ledger, trial_balance, chart_of_accounts)
            audit_record: Optional audit record for logging AI decisions
        
        Returns:
            Normalized data object (GeneralLedger, TrialBalance, or ChartOfAccounts)
        """
        logger.info(f"[parse_file] Parsing {filename} as {file_type}")
        
        extension = filename.lower().split(".")[-1]
        
        # Read file into DataFrame
        if extension == "csv":
            df = pd.read_csv(io.BytesIO(content))
        elif extension in ["xlsx", "xls"]:
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise ValueError(f"Unsupported file format: {extension}")
        
        if audit_record:
            audit_record.add_reasoning_step("File loaded for normalization", {
                "filename": filename,
                "file_type": file_type,
                "rows": len(df),
                "columns": list(df.columns)
            })
        
        # Route to appropriate normalizer
        if file_type == "general_ledger":
            return await self._normalize_gl(df, filename, audit_record)
        elif file_type == "trial_balance":
            return await self._normalize_tb(df, filename, audit_record)
        elif file_type == "chart_of_accounts":
            return await self._normalize_coa(df, filename, audit_record)
        else:
            raise ValueError(f"Unknown file type: {file_type}")
    
    # =========================================================================
    # Smart Upload  (multi-sheet Excel, CSV, chunked large files)
    # =========================================================================

    async def parse_upload(
        self,
        content: bytes,
        filename: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> dict:
        """
        Smart entry-point used by the upload-smart endpoint.

        * Handles .xlsx / .xls (multi-sheet) and .csv
        * Only sends a small sample to Gemini for column detection
        * Processes all rows locally, chunked for large files

        Returns:
            {"gl": GeneralLedger | None,
             "coa": ChartOfAccounts | None,
             "tb": TrialBalance | None,
             "detected_industry": str | None,
             "detected_basis": str | None}
        """
        extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        if extension in ("xlsx", "xls"):
            return await self._parse_excel_smart(content, filename, audit_record)
        elif extension == "csv":
            return await self._parse_csv_smart(content, filename, audit_record)
        else:
            raise ValueError(f"Unsupported file format: .{extension}")

    # -- Excel (multi-sheet) -------------------------------------------------

    async def _parse_excel_smart(
        self,
        content: bytes,
        filename: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> dict:
        """Handle .xlsx/.xls - detect sheets, classify, parse each."""
        file_bytes = io.BytesIO(content)

        # 1. Lightweight probe: sheet names + row counts via openpyxl read-only
        file_bytes.seek(0)
        wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        sheet_row_counts: dict[str, int] = {}
        for name in sheet_names:
            ws = wb[name]
            sheet_row_counts[name] = (ws.max_row or 1) - 1  # exclude header
        wb.close()

        total_rows = sum(sheet_row_counts.values())
        if total_rows > MAX_ROWS_TOTAL:
            raise ValueError(
                f"File too large: ~{total_rows:,} rows across {len(sheet_names)} sheets. "
                f"Maximum supported is {MAX_ROWS_TOTAL:,} rows."
            )

        # 2. Sample each sheet (first SAMPLE_ROWS rows only)
        file_bytes.seek(0)
        sheet_samples: dict[str, pd.DataFrame] = pd.read_excel(
            file_bytes, sheet_name=None, nrows=SAMPLE_ROWS
        )

        if audit_record:
            audit_record.add_reasoning_step("Excel file opened for smart parsing", {
                "filename": filename,
                "sheets": {n: {"rows": sheet_row_counts.get(n, 0),
                               "columns": list(df.columns)}
                           for n, df in sheet_samples.items()},
                "total_rows": total_rows,
            })

        logger.info(
            f"[_parse_excel_smart] {filename}: {len(sheet_names)} sheet(s), "
            f"~{total_rows:,} total rows"
        )

        # 3. Classify sheets
        if len(sheet_names) == 1:
            # Single sheet -> treat as GL
            classifications = {"general_ledger": sheet_names[0]}
        else:
            classifications = await self._classify_sheets(
                sheet_samples, sheet_row_counts, filename, audit_record
            )

        # 4. Parse each classified sheet
        result: dict = {
            "gl": None, "coa": None, "tb": None,
            "detected_industry": None, "detected_basis": None,
        }

        for data_type, sheet_name in classifications.items():
            if not sheet_name or sheet_name not in sheet_samples:
                continue

            rows_in_sheet = sheet_row_counts.get(sheet_name, 0)
            label = f"{filename}[{sheet_name}]"
            logger.info(
                f"[_parse_excel_smart] Parsing sheet '{sheet_name}' as "
                f"{data_type} ({rows_in_sheet:,} rows)"
            )

            if data_type == "general_ledger":
                if rows_in_sheet > CHUNK_SIZE:
                    result["gl"] = await self._read_gl_chunked(
                        content, sheet_name, label, audit_record
                    )
                else:
                    file_bytes.seek(0)
                    df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name)
                    result["gl"] = await self._normalize_gl(df, label, audit_record)

            elif data_type == "trial_balance":
                file_bytes = io.BytesIO(content)
                df = pd.read_excel(file_bytes, sheet_name=sheet_name)
                result["tb"] = await self._normalize_tb(df, label, audit_record)

            elif data_type == "chart_of_accounts":
                file_bytes = io.BytesIO(content)
                df = pd.read_excel(file_bytes, sheet_name=sheet_name)
                result["coa"] = await self._normalize_coa(df, label, audit_record)

        return result

    async def _classify_sheets(
        self,
        sheet_samples: dict[str, pd.DataFrame],
        sheet_row_counts: dict[str, int],
        filename: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> dict[str, str]:
        """Use AI to classify which Excel sheet is GL / COA / TB."""

        # Build a compact summary for Gemini
        sheets_desc_parts = []
        for name, df in sheet_samples.items():
            preview = df.head(5).to_csv(index=False)
            sheets_desc_parts.append(
                f"SHEET: \"{name}\"  (approx {sheet_row_counts.get(name, '?')} rows)\n"
                f"Columns: {list(df.columns)}\n"
                f"Sample:\n{preview}"
            )
        sheets_desc = "\n---\n".join(sheets_desc_parts)

        prompt = f"""You are a financial data analyst. An Excel workbook has been uploaded
with the following sheets. Classify each sheet as one of:
- general_ledger  (transaction journal / GL entries with dates, debits, credits)
- trial_balance   (account balances summary)
- chart_of_accounts  (list of accounts with codes and types)
- unknown         (not financial data)

FILE: {filename}

{sheets_desc}

Return JSON ONLY:
{{
    "classifications": {{
        "general_ledger": "sheet_name_or_null",
        "trial_balance": "sheet_name_or_null",
        "chart_of_accounts": "sheet_name_or_null"
    }},
    "reasoning": "brief explanation"
}}
"""

        result = await self.gemini.generate_json(
            prompt=prompt, purpose="sheet_classification"
        )

        if audit_record and result.get("audit"):
            audit_record.add_gemini_interaction(result["audit"])

        parsed = result.get("parsed", {})
        classifications_raw = parsed.get("classifications", {})

        # Normalise: only keep entries whose value is a real sheet name
        valid_names = set(sheet_samples.keys())
        classifications = {}
        for dtype in ("general_ledger", "trial_balance", "chart_of_accounts"):
            val = classifications_raw.get(dtype)
            if val and val in valid_names:
                classifications[dtype] = val

        if audit_record:
            audit_record.add_reasoning_step("AI classified Excel sheets", {
                "classifications": classifications,
                "reasoning": parsed.get("reasoning", ""),
            })

        # If AI found nothing usable, fall back: largest sheet = GL
        if not classifications:
            largest = max(sheet_row_counts, key=sheet_row_counts.get)  # type: ignore[arg-type]
            classifications["general_ledger"] = largest
            if audit_record:
                audit_record.add_reasoning_step(
                    "Sheet classification fallback: largest sheet treated as GL",
                    {"sheet": largest},
                )

        logger.info(f"[_classify_sheets] Classifications: {classifications}")
        return classifications

    # -- Chunked GL reading --------------------------------------------------

    async def _read_gl_chunked(
        self,
        content: bytes,
        sheet_name: str,
        label: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> GeneralLedger:
        """
        Read a large GL sheet in chunks using openpyxl streaming.

        1. Sample first SAMPLE_ROWS rows for AI column detection.
        2. Stream all rows through the detected mapping in CHUNK_SIZE batches.
        """
        logger.info(f"[_read_gl_chunked] Starting chunked read for '{label}'")

        # Step 1 - AI column detection on a small sample
        sample_df = pd.read_excel(
            io.BytesIO(content), sheet_name=sheet_name, nrows=SAMPLE_ROWS
        )
        mapping_result = await self._ai_detect_gl_columns(sample_df, label, audit_record)

        column_mapping = mapping_result.get("column_mapping", {})
        parsed_config = mapping_result  # date_format, currency_symbol, etc.

        if not column_mapping.get("date") or not column_mapping.get("account_code"):
            # Fallback to heuristic if AI fails
            column_mapping = await self._heuristic_detect_columns(
                sample_df, "general_ledger"
            )
            parsed_config = {}

        # Step 2 - Stream rows with openpyxl read-only
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        ws = wb[sheet_name]

        # Build header-index map from first row
        header_row = None
        header_map: dict[str, int] = {}  # column_name -> col_index
        entries: list[JournalEntry] = []
        chunk_num = 0
        rows_in_chunk = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                header_row = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                header_map = {name: idx for idx, name in enumerate(header_row)}
                continue

            # Build a pseudo-Series so helpers work
            row_data = {}
            for col_name, col_idx in header_map.items():
                row_data[col_name] = row[col_idx] if col_idx < len(row) else None
            row_series = pd.Series(row_data)

            try:
                entry_id = self._safe_get(
                    row_series, column_mapping.get("entry_id"),
                    f"GL-{row_idx:06d}"
                )
                date_val = self._safe_get(row_series, column_mapping.get("date"), "")
                account_code = self._safe_get(
                    row_series, column_mapping.get("account_code"), "0000"
                )
                account_name = self._safe_get(
                    row_series, column_mapping.get("account_name"), ""
                )
                debit = self._parse_amount(
                    row_series, column_mapping.get("debit"), parsed_config
                )
                credit = self._parse_amount(
                    row_series, column_mapping.get("credit"), parsed_config
                )
                description = self._safe_get(
                    row_series, column_mapping.get("description"), ""
                )
                vendor = self._safe_get(
                    row_series, column_mapping.get("vendor_or_customer"), None
                )

                date_str = self._normalize_date(
                    date_val, parsed_config.get("date_format")
                )

                entries.append(JournalEntry(
                    entry_id=str(entry_id),
                    date=date_str,
                    account_code=str(account_code),
                    account_name=str(account_name),
                    debit=debit,
                    credit=credit,
                    description=str(description),
                    vendor_or_customer=str(vendor) if vendor else None,
                ))
            except Exception as e:
                logger.warning(f"[_read_gl_chunked] Skip row {row_idx}: {e}")

            rows_in_chunk += 1
            if rows_in_chunk >= CHUNK_SIZE:
                chunk_num += 1
                logger.info(
                    f"[_read_gl_chunked] Chunk {chunk_num} done "
                    f"({len(entries):,} entries so far)"
                )
                if audit_record:
                    audit_record.add_reasoning_step(
                        f"Processed chunk {chunk_num}", {"entries_so_far": len(entries)}
                    )
                rows_in_chunk = 0

        wb.close()

        if audit_record:
            audit_record.add_reasoning_step(
                f"Chunked GL parsing complete for '{label}'",
                {"total_entries": len(entries), "chunks_processed": chunk_num + 1},
            )

        dates = [e.date for e in entries if e.date]
        period_start = min(dates) if dates else ""
        period_end = max(dates) if dates else ""

        logger.info(f"[_read_gl_chunked] Done: {len(entries):,} entries from '{label}'")
        return GeneralLedger(
            company_id="uploaded",
            entries=entries,
            period_start=period_start,
            period_end=period_end,
        )

    async def _ai_detect_gl_columns(
        self,
        sample_df: pd.DataFrame,
        label: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> dict:
        """
        Run AI column detection on a sample DataFrame and return the
        parsed config dict (column_mapping, date_format, etc.).
        Does NOT build entries -- only detects the mapping.
        """
        sample_rows = sample_df.head(SAMPLE_ROWS).to_csv(index=False)
        all_columns = list(sample_df.columns)

        prompt = f"""You are a financial data parser. Analyze this General Ledger file and identify the column mappings.

FILE NAME: {label}
COLUMNS FOUND: {all_columns}

SAMPLE DATA (first rows):
{sample_rows}

TASK: Identify which columns correspond to these required fields:
- entry_id: Transaction/entry identifier
- date: Transaction date
- account_code: Account number/code
- account_name: Account name/description
- debit: Debit amount
- credit: Credit amount
- description: Transaction description/memo
- vendor_or_customer: Vendor or customer name (optional)

Also detect:
- date_format: The format of dates (e.g., "YYYY-MM-DD", "MM/DD/YYYY")
- currency_symbol: If amounts have currency symbols
- has_thousands_separator: If amounts use commas for thousands

RETURN JSON ONLY:
{{
    "column_mapping": {{
        "entry_id": "actual_column_name_or_null",
        "date": "actual_column_name",
        "account_code": "actual_column_name",
        "account_name": "actual_column_name_or_null",
        "debit": "actual_column_name",
        "credit": "actual_column_name",
        "description": "actual_column_name_or_null",
        "vendor_or_customer": "actual_column_name_or_null"
    }},
    "date_format": "detected format",
    "currency_symbol": "$" or null,
    "has_thousands_separator": true/false,
    "parsing_notes": ["any observations about the data"]
}}
"""

        result = await self.gemini.generate_json(
            prompt=prompt, purpose="gl_column_detection"
        )

        if audit_record and result.get("audit"):
            audit_record.add_gemini_interaction(result["audit"])

        if result.get("error"):
            logger.warning(f"[_ai_detect_gl_columns] AI failed: {result.get('error')}")
            if audit_record:
                audit_record.add_reasoning_step(
                    "AI column detection failed (chunked path)", {"error": result.get("error")}
                )
            return {}

        parsed = result.get("parsed", {})

        if audit_record:
            audit_record.add_reasoning_step("AI detected column mappings (chunked path)", {
                "mapping": parsed.get("column_mapping", {}),
                "parsing_notes": parsed.get("parsing_notes", []),
            })

        return parsed

    # -- CSV smart parsing ---------------------------------------------------

    async def _parse_csv_smart(
        self,
        content: bytes,
        filename: str,
        audit_record: Optional[AuditRecord] = None,
    ) -> dict:
        """Parse a CSV file, with chunked support for large files."""
        file_bytes = io.BytesIO(content)

        # Quick row count estimate (count newlines)
        line_count = content.count(b"\n")
        if line_count > MAX_ROWS_TOTAL:
            raise ValueError(
                f"CSV too large: ~{line_count:,} rows. "
                f"Maximum supported is {MAX_ROWS_TOTAL:,} rows."
            )

        if audit_record:
            audit_record.add_reasoning_step("CSV file opened for smart parsing", {
                "filename": filename,
                "approx_rows": line_count,
            })

        logger.info(f"[_parse_csv_smart] {filename}: ~{line_count:,} rows")

        if line_count <= CHUNK_SIZE:
            # Small file: read all at once, delegate to existing normalizer
            file_bytes.seek(0)
            df = pd.read_csv(file_bytes)
            gl = await self._normalize_gl(df, filename, audit_record)
            return {
                "gl": gl, "coa": None, "tb": None,
                "detected_industry": None, "detected_basis": None,
            }

        # Large CSV: chunked processing
        # Step 1 - AI detection on sample
        file_bytes.seek(0)
        sample_df = pd.read_csv(file_bytes, nrows=SAMPLE_ROWS)
        mapping_result = await self._ai_detect_gl_columns(sample_df, filename, audit_record)

        column_mapping = mapping_result.get("column_mapping", {})
        parsed_config = mapping_result

        if not column_mapping.get("date") or not column_mapping.get("account_code"):
            column_mapping = await self._heuristic_detect_columns(
                sample_df, "general_ledger"
            )
            parsed_config = {}

        # Step 2 - Read in chunks
        entries: list[JournalEntry] = []
        file_bytes.seek(0)
        chunk_num = 0
        entry_counter = 0

        for chunk_df in pd.read_csv(file_bytes, chunksize=CHUNK_SIZE):
            chunk_num += 1
            for _, row in chunk_df.iterrows():
                try:
                    entry_id = self._safe_get(
                        row, column_mapping.get("entry_id"), f"GL-{entry_counter:06d}"
                    )
                    date_val = self._safe_get(row, column_mapping.get("date"), "")
                    account_code = self._safe_get(
                        row, column_mapping.get("account_code"), "0000"
                    )
                    account_name = self._safe_get(
                        row, column_mapping.get("account_name"), ""
                    )
                    debit = self._parse_amount(
                        row, column_mapping.get("debit"), parsed_config
                    )
                    credit = self._parse_amount(
                        row, column_mapping.get("credit"), parsed_config
                    )
                    description = self._safe_get(
                        row, column_mapping.get("description"), ""
                    )
                    vendor = self._safe_get(
                        row, column_mapping.get("vendor_or_customer"), None
                    )
                    date_str = self._normalize_date(
                        date_val, parsed_config.get("date_format")
                    )

                    entries.append(JournalEntry(
                        entry_id=str(entry_id),
                        date=date_str,
                        account_code=str(account_code),
                        account_name=str(account_name),
                        debit=debit,
                        credit=credit,
                        description=str(description),
                        vendor_or_customer=str(vendor) if vendor else None,
                    ))
                    entry_counter += 1
                except Exception as e:
                    logger.warning(f"[_parse_csv_smart] Skip row: {e}")
                    entry_counter += 1

            logger.info(
                f"[_parse_csv_smart] Chunk {chunk_num} done "
                f"({len(entries):,} entries so far)"
            )
            if audit_record:
                audit_record.add_reasoning_step(
                    f"CSV chunk {chunk_num} processed",
                    {"entries_so_far": len(entries)},
                )

        if audit_record:
            audit_record.add_reasoning_step("CSV chunked parsing complete", {
                "total_entries": len(entries),
                "chunks": chunk_num,
            })

        dates = [e.date for e in entries if e.date]
        period_start = min(dates) if dates else ""
        period_end = max(dates) if dates else ""

        gl = GeneralLedger(
            company_id="uploaded",
            entries=entries,
            period_start=period_start,
            period_end=period_end,
        )
        return {
            "gl": gl, "coa": None, "tb": None,
            "detected_industry": None, "detected_basis": None,
        }

    # =========================================================================
    # Existing normalizers (unchanged)
    # =========================================================================

    async def _normalize_gl(
        self, 
        df: pd.DataFrame, 
        filename: str,
        audit_record: Optional[AuditRecord] = None
    ) -> GeneralLedger:
        """Normalize General Ledger data using AI."""
        logger.info(f"[_normalize_gl] Normalizing GL from {filename}")
        
        # Try AI-powered parsing first
        result = await self._ai_parse_gl(df, filename, audit_record)
        
        if result:
            return result
        
        # Fallback to heuristic parsing
        logger.info("[_normalize_gl] AI parsing failed, using heuristic fallback")
        if audit_record:
            audit_record.add_reasoning_step("Using heuristic fallback for GL parsing")
        
        return await self._heuristic_parse_gl(df)
    
    async def _ai_parse_gl(
        self,
        df: pd.DataFrame,
        filename: str,
        audit_record: Optional[AuditRecord] = None
    ) -> Optional[GeneralLedger]:
        """Use Gemini AI to parse General Ledger data."""
        
        # Prepare sample data for AI
        sample_rows = df.head(10).to_csv(index=False)
        all_columns = list(df.columns)
        
        prompt = f"""You are a financial data parser. Analyze this General Ledger file and identify the column mappings.

FILE NAME: {filename}
COLUMNS FOUND: {all_columns}

SAMPLE DATA (first 10 rows):
{sample_rows}

TASK: Identify which columns correspond to these required fields:
- entry_id: Transaction/entry identifier
- date: Transaction date
- account_code: Account number/code
- account_name: Account name/description
- debit: Debit amount
- credit: Credit amount
- description: Transaction description/memo
- vendor_or_customer: Vendor or customer name (optional)

Also detect:
- date_format: The format of dates (e.g., "YYYY-MM-DD", "MM/DD/YYYY")
- currency_symbol: If amounts have currency symbols
- has_thousands_separator: If amounts use commas for thousands

RETURN JSON ONLY:
{{
    "column_mapping": {{
        "entry_id": "actual_column_name_or_null",
        "date": "actual_column_name",
        "account_code": "actual_column_name",
        "account_name": "actual_column_name_or_null",
        "debit": "actual_column_name",
        "credit": "actual_column_name",
        "description": "actual_column_name_or_null",
        "vendor_or_customer": "actual_column_name_or_null"
    }},
    "date_format": "detected format",
    "currency_symbol": "$" or null,
    "has_thousands_separator": true/false,
    "parsing_notes": ["any observations about the data"]
}}
"""
        
        result = await self.gemini.generate_json(
            prompt=prompt,
            purpose="gl_column_detection"
        )
        
        if audit_record and result.get("audit"):
            audit_record.add_gemini_interaction(result["audit"])
        
        if result.get("error"):
            logger.warning(f"[_ai_parse_gl] AI parsing failed: {result.get('error')}")
            if audit_record:
                audit_record.add_reasoning_step("AI column detection failed", {
                    "error": result.get("error")
                })
            return None
        
        parsed = result.get("parsed", {})
        column_mapping = parsed.get("column_mapping", {})
        
        if not column_mapping.get("date") or not column_mapping.get("account_code"):
            logger.warning("[_ai_parse_gl] AI could not detect required columns")
            return None
        
        if audit_record:
            audit_record.add_reasoning_step("AI detected column mappings", {
                "mapping": column_mapping,
                "parsing_notes": parsed.get("parsing_notes", [])
            })
        
        # Parse entries using AI-detected mapping
        entries = []
        for row_num, (idx, row) in enumerate(df.iterrows()):
            try:
                # Get values using detected mapping
                entry_id = self._safe_get(row, column_mapping.get("entry_id"), f"GL-{row_num:04d}")
                date_val = self._safe_get(row, column_mapping.get("date"), "")
                account_code = self._safe_get(row, column_mapping.get("account_code"), "0000")
                account_name = self._safe_get(row, column_mapping.get("account_name"), "")
                debit = self._parse_amount(row, column_mapping.get("debit"), parsed)
                credit = self._parse_amount(row, column_mapping.get("credit"), parsed)
                description = self._safe_get(row, column_mapping.get("description"), "")
                vendor = self._safe_get(row, column_mapping.get("vendor_or_customer"), None)
                
                # Normalize date
                date_str = self._normalize_date(date_val, parsed.get("date_format"))
                
                entry = JournalEntry(
                    entry_id=str(entry_id),
                    date=date_str,
                    account_code=str(account_code),
                    account_name=str(account_name),
                    debit=debit,
                    credit=credit,
                    description=str(description),
                    vendor_or_customer=str(vendor) if vendor else None
                )
                entries.append(entry)
                
            except Exception as e:
                logger.warning(f"[_ai_parse_gl] Error parsing row {idx}: {e}")
        
        if audit_record:
            audit_record.add_reasoning_step(f"AI parsed {len(entries)} GL entries", {
                "entries_count": len(entries),
                "total_debits": sum(e.debit for e in entries),
                "total_credits": sum(e.credit for e in entries)
            })
        
        # Determine period from dates
        dates = [e.date for e in entries if e.date]
        period_start = min(dates) if dates else ""
        period_end = max(dates) if dates else ""
        
        logger.info(f"[_ai_parse_gl] AI parsed {len(entries)} entries")
        
        return GeneralLedger(
            company_id="uploaded",
            entries=entries,
            period_start=period_start,
            period_end=period_end
        )
    
    async def _normalize_tb(
        self, 
        df: pd.DataFrame,
        filename: str,
        audit_record: Optional[AuditRecord] = None
    ) -> TrialBalance:
        """Normalize Trial Balance data using AI."""
        logger.info(f"[_normalize_tb] Normalizing TB from {filename}")
        
        # Prepare sample for AI
        sample_rows = df.head(10).to_csv(index=False)
        all_columns = list(df.columns)
        
        prompt = f"""Analyze this Trial Balance file and identify column mappings.

FILE NAME: {filename}
COLUMNS: {all_columns}

SAMPLE DATA:
{sample_rows}

IDENTIFY:
- account_code: Account number
- account_name: Account name
- debit: Debit balance
- credit: Credit balance

RETURN JSON:
{{
    "column_mapping": {{
        "account_code": "column_name",
        "account_name": "column_name_or_null",
        "debit": "column_name",
        "credit": "column_name"
    }},
    "has_thousands_separator": true/false,
    "currency_symbol": "$" or null
}}
"""
        
        result = await self.gemini.generate_json(prompt=prompt, purpose="tb_column_detection")
        
        if audit_record and result.get("audit"):
            audit_record.add_gemini_interaction(result["audit"])
        
        # Use AI mapping or fall back to heuristics
        if result.get("parsed"):
            column_mapping = result["parsed"].get("column_mapping", {})
            parsed_config = result.get("parsed", {})
        else:
            column_mapping = await self._heuristic_detect_columns(df, "trial_balance")
            parsed_config = {}
        
        if audit_record:
            audit_record.add_reasoning_step("TB column mapping detected", {
                "mapping": column_mapping,
                "ai_powered": bool(result.get("parsed"))
            })
        
        rows = []
        total_debits = 0.0
        total_credits = 0.0
        
        for _, row in df.iterrows():
            debit = self._parse_amount(row, column_mapping.get("debit"), parsed_config)
            credit = self._parse_amount(row, column_mapping.get("credit"), parsed_config)
            
            tb_row = TrialBalanceRow(
                account_code=str(self._safe_get(row, column_mapping.get("account_code"), "")),
                account_name=str(self._safe_get(row, column_mapping.get("account_name"), "")),
                debit=debit,
                credit=credit,
                ending_balance=debit - credit
            )
            rows.append(tb_row)
            total_debits += debit
            total_credits += credit
        
        if audit_record:
            audit_record.add_reasoning_step(f"Parsed {len(rows)} TB rows", {
                "total_debits": total_debits,
                "total_credits": total_credits,
                "is_balanced": abs(total_debits - total_credits) < 0.01
            })
        
        return TrialBalance(
            company_id="uploaded",
            period_end=datetime.now().strftime("%Y-%m-%d"),
            rows=rows,
            total_debits=total_debits,
            total_credits=total_credits,
            is_balanced=abs(total_debits - total_credits) < 0.01
        )
    
    async def _normalize_coa(
        self, 
        df: pd.DataFrame,
        filename: str,
        audit_record: Optional[AuditRecord] = None
    ) -> ChartOfAccounts:
        """Normalize Chart of Accounts data using AI."""
        logger.info(f"[_normalize_coa] Normalizing COA from {filename}")
        
        sample_rows = df.head(10).to_csv(index=False)
        all_columns = list(df.columns)
        
        prompt = f"""Analyze this Chart of Accounts file and identify column mappings.

FILE NAME: {filename}
COLUMNS: {all_columns}

SAMPLE DATA:
{sample_rows}

IDENTIFY:
- code: Account code/number
- name: Account name
- type: Account type (asset, liability, equity, revenue, expense)

RETURN JSON:
{{
    "column_mapping": {{
        "code": "column_name",
        "name": "column_name",
        "type": "column_name_or_null"
    }},
    "detected_types": ["list of account types found"]
}}
"""
        
        result = await self.gemini.generate_json(prompt=prompt, purpose="coa_column_detection")
        
        if audit_record and result.get("audit"):
            audit_record.add_gemini_interaction(result["audit"])
        
        if result.get("parsed"):
            column_mapping = result["parsed"].get("column_mapping", {})
        else:
            column_mapping = await self._heuristic_detect_columns(df, "chart_of_accounts")
        
        if audit_record:
            audit_record.add_reasoning_step("COA column mapping detected", {
                "mapping": column_mapping,
                "ai_powered": bool(result.get("parsed"))
            })
        
        accounts = []
        for _, row in df.iterrows():
            code = str(self._safe_get(row, column_mapping.get("code"), ""))
            name = str(self._safe_get(row, column_mapping.get("name"), ""))
            account_type = str(self._safe_get(row, column_mapping.get("type"), "expense")).lower()
            
            # Infer account type from code if not provided
            if not column_mapping.get("type"):
                account_type = self._infer_account_type(code)
            
            # Infer normal balance from type
            normal_balance = "credit" if account_type in ["liability", "equity", "revenue"] else "debit"
            
            account = Account(
                code=code,
                name=name,
                type=account_type,
                normal_balance=normal_balance
            )
            accounts.append(account)
        
        if audit_record:
            audit_record.add_reasoning_step(f"Parsed {len(accounts)} COA accounts", {
                "types_found": list(set(a.type for a in accounts))
            })
        
        return ChartOfAccounts(
            company_id="uploaded",
            accounts=accounts
        )
    
    # =========================================================================
    # Helper Methods
    # =========================================================================
    
    def _safe_get(self, row, column: Optional[str], default):
        """Safely get a value from a row."""
        if not column or column not in row.index:
            return default
        val = row.get(column, default)
        if pd.isna(val):
            return default
        return val
    
    def _parse_amount(self, row, column: Optional[str], config: dict) -> float:
        """Parse an amount value, handling currency symbols and separators."""
        if not column:
            return 0.0
        
        val = row.get(column, 0)
        if pd.isna(val):
            return 0.0
        
        if isinstance(val, (int, float)):
            return float(val)
        
        # String processing
        val_str = str(val)
        
        # Remove currency symbols
        currency = config.get("currency_symbol", "$")
        if currency:
            val_str = val_str.replace(currency, "")
        
        # Remove common currency symbols
        for sym in ["$", "EUR", "GBP", "USD"]:
            val_str = val_str.replace(sym, "")
        
        # Remove thousands separators
        val_str = val_str.replace(",", "")
        
        # Handle parentheses as negative
        if val_str.startswith("(") and val_str.endswith(")"):
            val_str = "-" + val_str[1:-1]
        
        try:
            return float(val_str.strip())
        except (ValueError, TypeError):
            return 0.0
    
    def _normalize_date(self, date_val, date_format: Optional[str] = None) -> str:
        """Normalize a date value to YYYY-MM-DD format."""
        if pd.isna(date_val) or date_val == "":
            return ""
        
        # Already in correct format
        if isinstance(date_val, str) and len(date_val) == 10 and date_val[4] == "-":
            return date_val
        
        # Handle pandas Timestamp
        if hasattr(date_val, 'strftime'):
            return date_val.strftime("%Y-%m-%d")
        
        # Try to parse string dates
        date_str = str(date_val)
        
        formats_to_try = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
            "%B %d, %Y",
            "%b %d, %Y",
        ]
        
        for fmt in formats_to_try:
            try:
                parsed = datetime.strptime(date_str, fmt)
                return parsed.strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        # Return as-is if nothing works
        return date_str
    
    def _infer_account_type(self, code: str) -> str:
        """Infer account type from account code."""
        try:
            code_num = int(code.replace("-", "").replace(".", "")[:4])
            if 1000 <= code_num < 2000:
                return "asset"
            elif 2000 <= code_num < 3000:
                return "liability"
            elif 3000 <= code_num < 4000:
                return "equity"
            elif 4000 <= code_num < 5000:
                return "revenue"
            else:
                return "expense"
        except (ValueError, TypeError):
            return "expense"
    
    async def _heuristic_parse_gl(self, df: pd.DataFrame) -> GeneralLedger:
        """Fallback heuristic-based GL parsing."""
        column_mapping = await self._heuristic_detect_columns(df, "general_ledger")
        
        entries = []
        for row_num, (idx, row) in enumerate(df.iterrows()):
            entry = JournalEntry(
                entry_id=str(self._safe_get(row, column_mapping.get("entry_id"), f"GL-{row_num:04d}")),
                date=str(self._safe_get(row, column_mapping.get("date"), "")),
                account_code=str(self._safe_get(row, column_mapping.get("account_code"), "")),
                account_name=str(self._safe_get(row, column_mapping.get("account_name"), "")),
                debit=self._parse_amount(row, column_mapping.get("debit"), {}),
                credit=self._parse_amount(row, column_mapping.get("credit"), {}),
                description=str(self._safe_get(row, column_mapping.get("description"), "")),
                vendor_or_customer=str(self._safe_get(row, column_mapping.get("vendor"), "")) or None
            )
            entries.append(entry)
        
        dates = [e.date for e in entries if e.date]
        period_start = min(dates) if dates else ""
        period_end = max(dates) if dates else ""
        
        return GeneralLedger(
            company_id="uploaded",
            entries=entries,
            period_start=period_start,
            period_end=period_end
        )
    
    async def _heuristic_detect_columns(self, df: pd.DataFrame, file_type: str) -> dict:
        """Detect column mappings using heuristic pattern matching."""
        columns = list(df.columns)
        
        patterns = {
            "general_ledger": {
                "entry_id": ["id", "entry_id", "transaction_id", "ref", "reference", "entry id"],
                "date": ["date", "trans_date", "transaction_date", "posting_date", "trans date"],
                "account_code": ["account", "account_code", "acct", "gl_account", "account code"],
                "account_name": ["account_name", "account_desc", "account name"],
                "debit": ["debit", "dr", "debit_amount", "debit amount"],
                "credit": ["credit", "cr", "credit_amount", "credit amount"],
                "description": ["memo", "description", "narrative", "details"],
                "vendor": ["vendor", "payee", "customer", "vendor_or_customer", "name"]
            },
            "trial_balance": {
                "account_code": ["account", "account_code", "acct", "account code"],
                "account_name": ["account_name", "description", "account name"],
                "debit": ["debit", "dr", "debit_balance", "debit balance"],
                "credit": ["credit", "cr", "credit_balance", "credit balance"]
            },
            "chart_of_accounts": {
                "code": ["code", "account_code", "account_number", "acct", "account code"],
                "name": ["name", "account_name", "description", "account name"],
                "type": ["type", "account_type", "category", "account type"]
            }
        }
        
        mapping = {}
        pattern_set = patterns.get(file_type, {})
        
        for field, possible_names in pattern_set.items():
            for col in columns:
                col_lower = col.lower().replace(" ", "_")
                if col_lower in possible_names or col.lower() in possible_names:
                    mapping[field] = col
                    break
        
        return mapping
